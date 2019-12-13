import numpy as np
import statsmodels.api as sm
import pickle
import os.path
import openpyxl
import xlrd, xlwt

fname_list_cpg = 'list_cpg.pickle'  # list cpg
fname_dict_cpg_chr = "dict_cpg_chr.pickle" # dictionary cpg-chromosome
fname_dict_cpg_gene = "dict_cpg_gene.pickle" # dictionary cpg-gene
fname_dict_cpg_num = "dict_cpg_num.pickle" # dictionary cpg-number
fname_matrix_betas = "matrix_betas.pickle"

# открыть файлы из
load_path = "C:/Users/PC/Documents/DNA/Linear_Regression"
load_path += "/table.xlsx"
# сохранить файл в
save_path = "C:/Users/PC/Documents"
save_path += '/table.xlsx'
def open_files(name): # создает .pkl которых нет

    cpg_key = 'ID_REF'
    chr_key = 'CHR'
    gene_key = 'UCSC_REFGENE_NAME'
    dict_cpg_gene_d = {}
    dict_cpg_chr_d = {}
    cpg_bad = []
    list_cpg = []

    bad_file = open('bad_cpgs.txt', 'r')
    for line in bad_file:
        cpg_bad.append(line.rstrip())

    bad_file.close()

    file = open('annotations.txt', 'r')
    line = file.readline().rstrip()
    line_list = line.rstrip().split('\t')
    cpg_id = line_list.index(cpg_key)
    chr_id = line_list.index(chr_key)
    gene_id = line_list.index(gene_key)
    i = 0
    for line in file:
        line_list = line.rstrip().split('\t')
        if line_list[chr_id] != 'X' and line_list[chr_id] != 'Y':
            if line_list[cpg_id] in cpg_bad:
                continue
            else:
                list_cpg.append(line_list[cpg_id])
                i+=1
                print(i)
                if name == fname_dict_cpg_chr:
                     dict_cpg_chr_d[line_list[cpg_id]] = line_list[chr_id]
                if name == fname_dict_cpg_gene:
                    dict_cpg_gene_d[line_list[cpg_id]] = line_list[gene_id]

    if name == fname_list_cpg:
        with open(fname_list_cpg, 'wb') as handle:
            pickle.dump(list_cpg, handle, protocol=pickle.HIGHEST_PROTOCOL)
        return list_cpg

    file.close()

    dict_cpg_num = {}
    num_rows = 0
    for cpg in list_cpg:
        dict_cpg_num[cpg] = num_rows
        num_rows += 1

    if name == fname_matrix_betas:
        file_betas = open('betas.txt', 'r')
        line = file_betas.readline()
        line_list = line.split('\t')

        num_cols = len(line_list) - 1  # кол-во столбцов

        matrix_betas = np.zeros((num_rows, num_cols), dtype=float)
        rows = 0
        for line in file_betas:
            line = line.replace('\n', '')
            line_list = line.split('\t')
            if line_list[0] in line_cpg:
                rows_betas = [float(s) for s in line_list[1::]]
                matrix_betas[rows, :] = rows_betas
                rows += 1
                print(rows)

        file_betas.close()

    if name == fname_list_cpg:
        with open(fname_list_cpg, 'wb') as handle:
            pickle.dump(list_cpg, handle, protocol=pickle.HIGHEST_PROTOCOL)
        return list_cpg
    if name == fname_dict_cpg_chr:
        with open(fname_dict_cpg_chr, 'wb') as handle:
            pickle.dump(dict_cpg_chr_d, handle, protocol=pickle.HIGHEST_PROTOCOL)
        return dict_cpg_chr_d
    if name == fname_dict_cpg_gene:
        with open(fname_dict_cpg_gene, 'wb') as handle:
            pickle.dump(dict_cpg_gene_d, handle, protocol=pickle.HIGHEST_PROTOCOL)
        return dict_cpg_gene_d
    if name == fname_dict_cpg_num:
        with open(fname_dict_cpg_num, 'wb') as handle:
            pickle.dump(dict_cpg_num, handle, protocol=pickle.HIGHEST_PROTOCOL)
        return dict_cpg_num
    if name == fname_matrix_betas:
        with open(fname_matrix_betas, 'wb') as handle:
            pickle.dump(matrix_betas, handle, protocol=pickle.HIGHEST_PROTOCOL)
        return matrix_betas


if os.path.isfile(fname_list_cpg):
    with open(fname_list_cpg, 'rb') as handle:
        line_cpg = pickle.load(handle)
else:
    line_cpg = open_files(fname_list_cpg)

if os.path.isfile(fname_dict_cpg_chr):
    with open(fname_dict_cpg_chr, 'rb') as handle:
        dict_cpg_chr = pickle.load(handle)
else:
    dict_cpg_chr = open_files(fname_dict_cpg_chr)

if os.path.isfile(fname_dict_cpg_gene):
    with open(fname_dict_cpg_gene, 'rb') as handle:
        dict_cpg_gene = pickle.load(handle)
else:
    dict_cpg_gene = open_files(fname_dict_cpg_gene)

if os.path.isfile(fname_dict_cpg_num):
    with open(fname_dict_cpg_num, 'rb') as handle:
        dict_cpg_num = pickle.load(handle)
else:
    dict_cpg_num = open_files(fname_dict_cpg_num)

if os.path.isfile(fname_matrix_betas):
    with open(fname_matrix_betas, 'rb') as handle:
        matrix_betas = pickle.load(handle)
else:
    matrix_betas = open_files(fname_matrix_betas)

p_age = {}
file = open('observables.txt', 'r')
age_key = 'age'
p_key = 'geo_accession'

line = file.readline().rstrip()
line_list = line.split('\t')
p_id = line_list.index(p_key)
age_id = line_list.index(age_key)
line_age = []
for line in file:
    line_list = line.rstrip().split('\t')

    p = line_list[p_id]
    age = line_list[age_id]
    p_age[p] = age
    line_age.append(int(age))
a = 0
#wb = xlwt.Workbook()
#ws = wb.add_sheet('Table')
#wb.save(save_path)
wb = openpyxl.load_workbook(filename = load_path)
ws = wb['Table']

data = ["СpG", "Gene", "Chromosome", "R-squared", "Adj.R-squared", "F-statistic", "Prob(F-statistic)", "Intercept", "Slope"]
i = 1
for x in data:
    ws.cell(row=1, column=i).value = x
    i += 1

k = 2 # по столбцам
j = 0 # по строчкам
i = 0
while i < len(matrix_betas):
    X = sm.add_constant(line_age)
    model = sm.OLS(matrix_betas[i], X)
    results = model.fit()
    ws.cell(row=k, column=1).value = line_cpg[i]
    ws.cell(row=k, column=2).value = str(dict_cpg_gene[line_cpg[i]])
    ws.cell(row=k, column=3).value = dict_cpg_chr[line_cpg[i]]
    ws.cell(row=k, column=4).value = results.rsquared
    ws.cell(row=k, column=5).value = results.rsquared_adj
    ws.cell(row=k, column=6).value = results.fvalue
    ws.cell(row=k, column=7).value = results.f_pvalue
    ws.cell(row=k, column=8).value = results.params[0]
    ws.cell(row=k, column=9).value = results.params[1]

    k += 1
    print(i)
    i += 1
wb.save(save_path)

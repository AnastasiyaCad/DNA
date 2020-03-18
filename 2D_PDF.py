import seaborn as sns
import matplotlib.pyplot as plt
import pickle
import math
import openpyxl
import xlrd, xlwt
import statsmodels.api as sm
import plotly
from plotly import tools
#import plotly.plotly as py
#import chart_studio.plotly as py
import plotly.graph_objs as go
import numpy as np
#import cmocean
import os

"""
load_path = "C:/Users/PC/Documents/DNA/Linear_Regression"
load_path += "/table_c.xlsx"
# сохранить файл в
save_path = "C:/Users/PC/Documents"
save_path += '/table_c.xlsx'
"""

fname_list_cpg = 'list_cpg.pickle'  # list cpg
fname_dict_cpg_chr = "dict_cpg_chr.pickle"  # dictionary cpg-chromosome
fname_dict_cpg_gene = "dict_cpg_gene.pickle"  # dictionary cpg-gene
fname_dict_cpg_num = "dict_cpg_num.pickle"  # dictionary cpg-number
fname_matrix_betas = "matrix_betas.pickle"

with open(fname_list_cpg, 'rb') as handle:
    line_cpg = pickle.load(handle)
with open(fname_dict_cpg_chr, 'rb') as handle:
    dict_cpg_chr = pickle.load(handle)
with open(fname_dict_cpg_gene, 'rb') as handle:
    dict_cpg_gene = pickle.load(handle)
with open(fname_dict_cpg_num, 'rb') as handle:
    dict_cpg_num = pickle.load(handle)
with open(fname_matrix_betas, 'rb') as handle:
    matrix_betas = pickle.load(handle)

persona_gender = {}
file = open('observables.txt', 'r')
age_key = 'age'
gender_key = 'gender'
p_key = 'geo_accession'

line = file.readline().rstrip()
line_list = line.split('\t')
p_id = line_list.index(p_key)
age_id = line_list.index(age_key)
gender_id = line_list.index(gender_key)
line_gender = []
line_age = []
for line in file:
    line_list = line.rstrip().split('\t')
    p = line_list[p_id]
    gender = line_list[gender_id]
    age = line_list[age_id]
    persona_gender[p] = gender
    line_gender.append(gender)
    line_age.append(int(age))

""""
подготовка таблицы
wb = openpyxl.load_workbook(filename = load_path)
ws = wb['Table']
data = ["СpG", "R-squared_f", "R-squared_m", "Intercept_f", "Intercept_m", "Slope_f", "Slope_m", "I"]
i = 1
k = 2 # по столбцам
j = 0 # по строчкам
for x in data:
    ws.cell(row=1, column=i).value = x
    i += 1
"""

#cpg_list = ["cg08037478", "cg04946709", "cg00167275", "cg00804338", "cg07381872", "cg23778841", "cg23719534", "cg24016844", "cg23256579", "cg06710937"]
cpg_list = ["cg08037478"]
for cpg in cpg_list:
    i = 0
    dict_cpg_num[cpg]

    line_age_male = []
    line_age_female = []
    matrix_betas_male = []
    matrix_betas_female = []
    i = 0
    while i < len(line_gender):
        if line_gender[i] == 'M':
            line_age_male.append(line_age[i])
            matrix_betas_male.append(matrix_betas[dict_cpg_num[cpg]][i])
        else:
            line_age_female.append(line_age[i])
            matrix_betas_female.append(matrix_betas[dict_cpg_num[cpg]][i])
        i += 1

    """
    plt.scatter(line_age_male, matrix_betas_male, label='', color='b', s=8)
    plt.scatter(line_age_female, matrix_betas_female, label='', color='r', s=8)
    name = cpg
    plt.title(name)
    plt.xlabel('age')
    plt.ylabel('β')
    plt.show()
    """

    age_num = []
    list_age_new = []
    line_age_sort = sorted(line_age)
    i = -1
    for age in line_age_sort:
        if age not in list_age_new:
            list_age_new.append(age)
            age_num.append(1)
            i += 1
        else:
            age_num[i] += 1

    bins_age = 30
    bins_betas = 30
    eps = 10 ** -10
    # matrix_male = [[0 for x in range(bins_age)] for y in range(bins_betas)]
    # matrix_female = [[0 for x in range(bins_age)] for y in range(bins_betas)]
    matrix_male = []
    matrix_female = []
    for i in range(bins_age):
        matrix_male.append([])
        matrix_female.append([])
        for j in range(bins_betas):
            matrix_male[i].append(0)
            matrix_female[i].append(0)

    x_min = min(line_age)
    x_max = max(line_age)

    y_min = min(matrix_betas[dict_cpg_num[cpg]])
    y_max = max(matrix_betas[dict_cpg_num[cpg]])

    dx = (x_max - x_min) / bins_age
    dy = (y_max - y_min) / bins_betas

    list_dx = []
    list_dy = []
    i = y_min + dy / 2
    while i < y_max:
        list_dy.append(i)
        i += dy

    j = x_min + dx / 2
    while j < x_max:
        list_dx.append(j)
        j += dx

    matrix_male_new = []
    matrix_female_new = []
    for i in range(bins_betas):
        matrix_male_new.append([])
        matrix_female_new.append([])
        for j in range(bins_age):
            matrix_male_new[i].append(0)
            matrix_female_new[i].append(0)

    i = 0
    for y in matrix_betas_male:
        x = line_age_male[i]
        i += 1

        IDx = math.floor(((x - x_min) * bins_age) / (x_max - x_min + eps))
        IDy = math.floor(((y - y_min) * bins_betas) / (y_max - y_min + eps))
        matrix_male[IDx][IDy] += 1
        matrix_male_new[IDy][IDx] += 1

    i = 0
    for y in matrix_betas_female:
        x = line_age_female[i]
        i += 1

        IDx = math.floor(((x - x_min) * bins_age) / (x_max - x_min + eps))
        IDy = math.floor(((y - y_min) * bins_betas) / (y_max - y_min + eps))
        matrix_female[IDx][IDy] += 1
        matrix_female_new[IDy][IDx] += 1

    I = 0
    Pi_line = []
    for ix in range(bins_age):
        for iy in range(bins_betas):
            Pi = min(matrix_male[ix][iy], matrix_female[ix][iy])
            Pi_line.append(Pi)
            I += Pi
    print(I)

    """
    # гистограмма общая
    plt.hist2d(line_age, matrix_betas[dict_cpg_num[cpg]], (bins_age, bins_betas), cmap=plt.cm.jet)
    name = cpg
    plt.title(name)
    plt.xlabel('age')
    plt.ylabel('β')
    plt.colorbar()
    plt.show()
    """

    # проба go
    fn = cpg
   # dense = cmocean_to_plotly(cmocean.cm.dense, I)
  #  py.iplot(colorscale_plot(colorscale=dense, title='Dense'))

    list_contour = [go.Contour(z=matrix_female_new, x=list_dx, y=list_dy, colorscale='bluered', opacity=0.3), go.Contour(z=matrix_male_new, x=list_dx, y=list_dy, colorscale='blugrn', opacity=0.3)]
    fig = go.Figure(list_contour)
    fig.show()
    plotly.offline.plot(fig, filename=fn + '.html', auto_open=False, show_link=True)
    #plotly.io.write_image(fig, fn + '.png')
    #plotly.io.write_image(fig, fn + '.pdf')
    """
    figure = plt.subplots()
    figure.scatter(line_age_female, matrix_betas_female, matrix_female)
    figure.scatter(line_age_male, matrix_betas_male, matrix_male)
    figure.legend()
    figure.show()
    """

    i = 0
"""
занесение данных в таблицу
    X_f = sm.add_constant(line_age_female)
    model_f = sm.OLS(matrix_betas_female, X_f)
    results_f = model_f.fit()

    X_m = sm.add_constant(line_age_male)
    model_m = sm.OLS(matrix_betas_male, X_m)
    results_m = model_m.fit()

    ws.cell(row=k, column=1).value = cpg
    ws.cell(row=k, column=2).value = results_f.rsquared
    ws.cell(row=k, column=3).value = results_m.rsquared
    ws.cell(row=k, column=4).value = results_f.params[0]
    ws.cell(row=k, column=5).value = results_m.params[0]
    ws.cell(row=k, column=6).value = results_f.params[1]
    ws.cell(row=k, column=7).value = results_m.params[1]
    ws.cell(row=k, column=8).value = I

    k += 1
    print(k)

wb.save(load_path)
"""
